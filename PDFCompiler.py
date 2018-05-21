import sys
from pdfminer import *
from pycallgraph import Config
from pycallgraph import PyCallGraph
from pycallgraph.output import GraphvizOutput
import slate
import os
from PyPDF2 import PdfFileMerger, PdfFileReader
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
from pprint import pprint
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.pdfpage import PDFPage
from cStringIO import StringIO
from Tkinter import *
import Tkinter
import tkFileDialog

dir_path = os.path.dirname(os.path.realpath(__file__))

foundNext = True

indexList = []

errorOutput = []

bids = []
aR = []
jobP = []
sGA = []
district = []
operating = []
types = [operating, sGA, district, jobP, aR, bids]
typeName = []

storageLocation, outputLocation, excelLocation, noneLocation = ("", "", "", "")

"""
storageLocation = "C:/Users/wsheung/PycharmProjects/reportComp/storage"

excelLocation = "C:/Users/wsheung/PycharmProjects/reportComp/excel"

outputLocation = "C:/Users/wsheung/PycharmProjects/reportComp/output"

noneLocation = "C:/Users/wsheung/PycharmProjects/reportComp/blanks"
"""


# Builds the framework for bookMarks
def buildBookmark():
    directory = []

    districtName = ''

    lastFoundDistrictIndex = 0

    for y in range(0, len(types[1])):  # longest column ,come back to this later and pass it down from import excel
        foundDistrict = False

        op = []
        sg = []
        dis = []
        job = []
        arrr = []
        bi = []
        listings = [op, sg, dis, job, arrr, bi]

        if movedToNextDistrict(y):
            foundDistrict = True

        if foundDistrict:
            if str(types[0][y]) != "None":
                districtName = types[0][y]
            for i in range(0, len(types)):
                listings[i].append(types[i][y])
        if not foundDistrict:
            for i in range(0, len(types)):
                if str(types[i][y]) != "None":
                    directory[lastFoundDistrictIndex - 1][1][i].append(
                        types[i][y])  # -1 because list index starts at 0

        if foundDistrict:
            directory.append([districtName, listings])
            lastFoundDistrictIndex = lastFoundDistrictIndex + 1

    # print directory

    # pprint(directory) enable this if you want to print /test the result of the directory

    return directory


def movedToNextDistrict(rowNumber):
    for x in range(0, len(types)):
        if str(types[x][rowNumber]) == "None":
            return False

    return True


def importExcelDoc():
    os.chdir(excelLocation)
    listOfFiles = getMasterList(excelLocation)
    if len(listOfFiles) != 1:
        print "Input folder contain more than one excel file"
    wb = load_workbook(listOfFiles[0])
    ws = wb.active
    longestColumnLength = 0
    cells = ws['A2': 'F100']

    for c1, c2, c3, c4, c5, c6 in cells:
        if (str(c1.value) != "None" or str(c2.value) != "None" or str(c3.value) != "None" or str(
                c4.value) != "None" or str(c5.value) != "None" or str(c6.value) != "None"):
            operating.append(str(c1.value))
            district.append(str(c3.value))
            sGA.append(str(c2.value))
            jobP.append(str(c4.value))
            aR.append(str(c5.value))
            bids.append(str(c6.value))
            longestColumnLength = longestColumnLength + 1
            # print("{0:50}{1:50}{2:50}{3:50}{4:50}{5:50}".format(c1.value, c2.value, c3.value, c4.value, c5.value, c6.value))

    names = ws['A1': 'F1']

    for c1, c2, c3, c4, c5, c6 in names:
        typeName.append(str(c1.value))
        typeName.append(str(c2.value))
        typeName.append(str(c3.value))
        typeName.append(str(c4.value))
        typeName.append(str(c5.value))
        typeName.append(str(c6.value))

    # print typeName
    result = buildBookmark()

    # for r1 in titles:
    # typeName.append(r1.value)

    # print(typeName)
    # list = tuple(ws['A2':'A34'])
    # ws['K3']=100
    wb.save(listOfFiles[0])

    return result


def validEntry(x):
    files = getMasterList(x)
    return len(files) < 2


def askForDirectory():
    # Asks the user to input directory manually to find areas of interest, input and output location
    root = Tkinter.Tk()
    global storageLocation
    global excelLocation
    global outputLocation
    global noneLocation

    storageLocation = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                                title="Please select the folder that solely contains source pdfs")
    excelLocation = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                              title="Please select the folder that solely contains the excel inputFile")
    while not validEntry(excelLocation):
        print "Invalid entry, press cancel to exit"
        excelLocation = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                                  title="Please select the folder that solely contains the excel inputFile")

    outputLocation = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                               title="Please select the folder you want to receive the output")

    noneLocation = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                             title="Please select the folder that contains blank sheet substitutes")
    while not validEntry(noneLocation):
        print "Invalid entry, press cancel to exit"
        noneLocation = tkFileDialog.askdirectory(parent=root, initialdir="/",
                                                 title="Please select the folder that solely contains the excel inputFile")

        # label = Label(root,
        # text="Thanks for using pdfComp for this month end report, please give this program sometime to load (~5 minutes)  - William S")
    # root.mainloop()
    return


def do():
    print ("Let's Begin!")
    askForDirectory()
    startTime = datetime.now()
    excelInput = importExcelDoc()
    # pprint(excelInput)  # --------------------------------------print beautifully
    masterList = getMasterList(storageLocation)
    overAllLength = len(masterList)
    print ("Processing  ", overAllLength, "  files in the storage folder")
    for x in range(0, len(masterList)):
        print '\n', 'processing document number', x
        print masterList[x]
        processPDF(masterList[x])
        progress = int(x * 100 / overAllLength / 2)

        drawProgressBar(progress, 50)
    # pprint(indexList)   # --------------------------------------print beautifully
    result = assignToSpot(excelInput)
    os.chdir(outputLocation)
    outputStream = file("MonthEndReportResult_OUTPUT.pdf", "wb")
    result.write(outputStream)
    print result
    print (startTime)
    print (datetime.now())
    drawProgressBar(100, 20)
    return


def getMasterList(sLoc):
    os.chdir(sLoc)
    pdfList = os.listdir(os.getcwd())
    # print (pdfList)
    return pdfList


def getCompanyName(doc, t):
    global foundNext
    for n in types[t]:
        if t == 5:
            if doc.find(n, 0, 250) > 0:
                # print doc.find(n, 0, 250)
                # print n
                return n

        if t != 5:
            if doc.find(n, 0, 4000) > 0:
                if doc.find("00030", 0, 4000) >= 0 and (
                            n == "MAR  "):  # Extra if statement to catch the two Marshalltown location between MYERS and HARLAN
                    return "MARHE"
                if doc.find("GSW", 0, 4000) >= 0 and (n == "Large Projects"):
                    return "GSW - Large Projects"
                if (doc[:4000].count("Harlan") > 1 and n == "7Harlan") and t == 2:
                    return "Harlan"
                return n
    # print 'Found Nothing'
    # print ("error")
    return
    # raise SystemExit


def getTypeName(doc):  # type of document
    global foundNext
    for tn in typeName:
        if typeName.index(tn) == 2:
            a, b = typeName[2].split("/")
            if doc.find(a, 0, 4000) >= 0:
                foundNext = True
                return 2
            if doc.find(b, 0, 4000) >= 0:
                foundNext = True
                return 2  # Can probably optimize this better than just brute forcing this with index 1
        if doc.find(tn, 0, 4000) >= 0:
            foundNext = True
            return typeName.index(tn)
        if not foundNext and typeName.index(tn) == 5:
            return 0
    foundNext = False

    # print ("error")
    return 0
    # raise SystemExit


def processPDF(p):
    # documents = openFile(p, 'rb')
    documents = convert_pdf_to_txt(p)
    global foundNext
    foundNext = True
    lastCompName = ""
    for x in range(0, len(documents)):
        # print(documents[x])
        type_name = getTypeName(documents[x])
        # print(foundNext)
        if foundNext:
            compName = getCompanyName(documents[x], type_name)
        if (x > 0 and not foundNext) and type_name == 0:
            compName = lastCompName

        lastCompName = compName
        indexList.append([type_name, compName])

        # print("done zo")
        # print(typeName,compName)
    return


def convert_pdf_to_txt(path):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = None
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = file(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos = set()

    leng = 0
    listin = []

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password, caching=caching,
                                  check_extractable=True):
        interpreter.process_page(page)
        textTemp = retstr.getvalue()
        text = textTemp[leng:]
        leng = len(textTemp)
        # print text
        listin.append(text)

    fp.close()
    device.close()
    retstr.close()

    return listin


def assignToSpot(index):
    errorList = []
    monthlyReport = PdfFileMerger()
    monthlyReportLength = 0
    DistrictResult = monthlyReport.addBookmark("District Packages", 0,
                                               parent=None)  # change this later if added more parent bookmarks (Pending Importance: Medium)
    DistrictName = None
    #  above this
    dex = index
    for x in range(0, len(dex)):
        # var = x[0]
        drawProgressBar(int(50 + x * 100 / len(dex) / 2), 50)
        doneWithOne = True
        for i in range(0, len(dex[x][1])):

            if i == 1 and doneWithOne:
                i = 2
            if i == 2 and not doneWithOne:
                i = 1
            for z in range(0, len(dex[x][1][i])):
                root = dex[x][1][i][z]
                foundFile = False
                firstPageOfWriteUpMerged = False
                firstPageOfTypeMerged = False
                if root == "7Harlan" and i == 2:  # Introduce Canada and WPE exceptions (Pending importance : High)
                    root = "Harlan"
                if root == "GSW - Large Projects":
                    root = "Large Projects"
                if root == "MARHE":
                    root = "MAR  "
                    # print root
                for s in range(0, len(indexList)):
                    if indexList[s][0] == i:
                        if indexList[s][1] == root:
                            # print indexList[s]
                            # print ("found it")
                            interestFile = getInterestFile(s, storageLocation)
                            if interestFile is not None:
                                fp = interestFile[1]
                                interestLoc = interestFile[0]
                                monthlyReport.append(fileobj=fp, pages=(interestLoc, interestLoc + 1),
                                                     import_bookmarks=False)
                                monthlyReportLength = monthlyReportLength + 1
                                foundFile = True

                                if i == 0 and z == 0 and not firstPageOfWriteUpMerged:
                                    DistrictName = monthlyReport.addBookmark(str(root), monthlyReportLength - 1,
                                                                             DistrictResult)
                                    firstPageOfWriteUpMerged = True
                                if DistrictName is not None:
                                    if i == 1 and z == 0 and not firstPageOfTypeMerged:
                                        monthlyReport.addBookmark("SG&A", monthlyReportLength - 1, DistrictName)
                                        firstPageOfTypeMerged = True
                                    if i == 2 and z == 0 and not firstPageOfTypeMerged:
                                        monthlyReport.addBookmark("District Package", monthlyReportLength - 1,
                                                                  DistrictName)
                                        firstPageOfTypeMerged = True
                                    if i == 3 and z == 0 and not firstPageOfTypeMerged:
                                        monthlyReport.addBookmark("Job Profitability", monthlyReportLength - 1,
                                                                  DistrictName)
                                        firstPageOfTypeMerged = True
                                    if i == 4 and z == 0 and not firstPageOfTypeMerged:
                                        monthlyReport.addBookmark("A/R", monthlyReportLength - 1, DistrictName)
                                        firstPageOfTypeMerged = True
                                    if i == 5 and z == 0 and not firstPageOfTypeMerged:
                                        monthlyReport.addBookmark("Bid Logs", monthlyReportLength - 1, DistrictName)
                                        firstPageOfTypeMerged = True
                                    if i == 0 and z == 0 and not firstPageOfTypeMerged:
                                        monthlyReport.addBookmark("Write Up", monthlyReportLength - 1, DistrictName)
                                        firstPageOfTypeMerged = True
                if not foundFile:
                    # print ("Can't find ", root)
                    if i == 5:  # This whole thing needs to be done in a for loop again, come back later :D (Pending Importance = Low)
                        noneFile = getInterestFile(5, noneLocation)

                        errorList.append(
                            ("Can't find ", root, " in category ", typeName[5], "at page ", monthlyReportLength))
                    if i == 3:
                        noneFile = getInterestFile(3, noneLocation)

                        errorList.append(
                            ("Can't find ", root, " in category ", typeName[3], "at page ", monthlyReportLength))
                    if i == 4:
                        noneFile = getInterestFile(4, noneLocation)

                        errorList.append(
                            ("Can't find ", root, " in category ", typeName[4], "at page ", monthlyReportLength))
                    if i == 2:
                        noneFile = getInterestFile(1, noneLocation)

                        errorList.append(
                            ("Can't find ", root, " in category ", typeName[1], "at page ", monthlyReportLength))
                    if i == 1:
                        noneFile = getInterestFile(2, noneLocation)

                        errorList.append(
                            ("Can't find ", root, " in category ", typeName[2], "at page ", monthlyReportLength))
                    if i == 0:
                        noneFile = getInterestFile(0, noneLocation)
                        errorList.append(
                            ("Can't find ", root, " in category ", typeName[0], "at page ", monthlyReportLength))

                    monthlyReportLength = monthlyReportLength + 1
                    fp2 = noneFile[1]
                    intLoc = noneFile[0]
                    monthlyReport.append(fp2, pages=(intLoc, intLoc + 1), import_bookmarks=False)
            if i != 0:
                doneWithOne = False
    for x in errorList:
        print "\n", x
    return monthlyReport


def getInterestFile(s1, path):
    os.chdir(path)
    targetIndex = s1
    currentIndex = 0
    interestList = getMasterList(path)
    targetFile = []
    for x in range(0, len(interestList)):
        fp = file(interestList[x], 'rb')
        read_pdf = PdfFileReader(fp)
        numPages = read_pdf.getNumPages()
        currentIndex = currentIndex + numPages
        if currentIndex > targetIndex:
            targetFile.append(targetIndex - currentIndex + numPages)
            targetFile.append(fp)
            return targetFile

    print ("error")  # should NEVER run beyond the line above
    return None


def drawProgressBar(percent, barLen=100):
    sys.stdout.write("\r")
    progress = ""
    for i in range(barLen):
        if (i * 100) < int(barLen * percent):
            progress += "="
        else:
            progress += " "
    sys.stdout.write("[ %s ] %.2f%%" % (progress, percent))
    sys.stdout.flush()


###########################################################################################
# No longer used, slower, clunkier, and less customization version of covert_pdf_to_text  #
###########################################################################################
def openFile(fileName, s):
    startTime = datetime.now()
    global foundNext
    foundNext = True
    with open(fileName, s) as f:
        doc = slate.PDF(f)
        # for x in range(0, len(doc)):               testing individual pages
        # if (doc[x].find("Bids by District", 0, 2000)>=0):
        # print("Found it!")
        # print (doc[0])
    print startTime
    print datetime.now()
    return doc


# ---------------------------------------TESTING AREA-----------------------------------------#
# openFile("77-Job Prof.pdf", 'rb')
# run()
# importExcelDoc()
# getMasterList()
# os.chdir(storageLocation)
# openFile("2.pdf", "rb")
do()
